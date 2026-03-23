import os
import tempfile
from groq import Groq
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes
import PyPDF2
import docx
import openpyxl
import base64

TELEGRAM_BOT_TOKEN = "8668267460:AAHWtFyxgUgsVcXRTXjVTrzBZyWIlAxGb2c"
GROQ_API_KEY = "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"

client = Groq(api_key=GROQ_API_KEY)
chat_histories = {}

SYSTEM_PROMPT = """Tu ek helpful AI assistant hai. 
Bilkul usi tarah baat kar jaise do dost baat karte hain - casual, friendly aur simple Hindi mein.
Jaise "kya scene hai", "bilkul bhai", "dekh yaar", "sahi hai" - is tarah ki natural Hindi use kar.
Agar user English mein pooche toh English mein jawab de, warna Hindi mein baat kar.
Kabhi bhi boring formal language mat use karna."""

def get_ai_response(user_id, user_message):
    if user_id not in chat_histories:
        chat_histories[user_id] = [{"role": "system", "content": SYSTEM_PROMPT}]
    chat_histories[user_id].append({"role": "user", "content": user_message})
    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=chat_histories[user_id]
    )
    ai_reply = response.choices[0].message.content
    chat_histories[user_id].append({"role": "assistant", "content": ai_reply})
    return ai_reply

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Kya scene hai yaar! 👋\nMain tera AI Assistant hoon 🤖\n\n"
        "Main ye sab kar sakta hoon:\n"
        "💬 Text messages\n"
        "🎤 Voice messages\n"
        "📄 PDF files\n"
        "🖼️ Images\n"
        "📝 Word documents\n"
        "📊 Excel files\n\n"
        "/reset - Chat clear karo"
    )

async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_histories[update.message.from_user.id] = []
    await update.message.reply_text("Done yaar! Nayi baat shuru karte hain 😊")

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    user_message = update.message.text
    await update.message.chat.send_action("typing")
    try:
        ai_reply = get_ai_response(user_id, user_message)
        await update.message.reply_text(ai_reply)
    except Exception as e:
        await update.message.reply_text(f"Yaar kuch masla aa gaya: {str(e)}")

async def handle_voice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    await update.message.chat.send_action("typing")
    try:
        voice = update.message.voice
        voice_file = await context.bot.get_file(voice.file_id)
        with tempfile.NamedTemporaryFile(suffix=".ogg", delete=False) as tmp:
            await voice_file.download_to_drive(tmp.name)
            tmp_path = tmp.name
        with open(tmp_path, "rb") as audio_file:
            transcription = client.audio.transcriptions.create(
                model="whisper-large-v3",
                file=audio_file
            )
        user_message = transcription.text
        os.unlink(tmp_path)

        await update.message.reply_text(f"🎤 Tune kaha: {user_message}")

        # Sirf text mein jawab — voice reply nahi!
        ai_reply = get_ai_response(user_id, user_message)
        await update.message.reply_text(ai_reply)

    except Exception as e:
        await update.message.reply_text(f"Yaar kuch masla aa gaya: {str(e)}")

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    await update.message.chat.send_action("typing")
    try:
        doc = update.message.document
        file = await context.bot.get_file(doc.file_id)
        file_name = doc.file_name.lower()

        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file_name)[1]) as tmp:
            await file.download_to_drive(tmp.name)
            tmp_path = tmp.name

        extracted_text = ""

        if file_name.endswith(".pdf"):
            await update.message.reply_text("📄 PDF parh raha hoon yaar...")
            with open(tmp_path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    extracted_text += page.extract_text() or ""

        elif file_name.endswith(".docx"):
            await update.message.reply_text("📝 Word file dekh raha hoon...")
            doc_file = docx.Document(tmp_path)
            for para in doc_file.paragraphs:
                extracted_text += para.text + "\n"

        elif file_name.endswith(".xlsx") or file_name.endswith(".xls"):
            await update.message.reply_text("📊 Excel dekh raha hoon...")
            wb = openpyxl.load_workbook(tmp_path)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                extracted_text += f"Sheet: {sheet}\n"
                for row in ws.iter_rows(values_only=True):
                    extracted_text += " | ".join([str(c) for c in row if c is not None]) + "\n"

        else:
            await update.message.reply_text("Yaar ye format support nahi hai!")
            os.unlink(tmp_path)
            return

        os.unlink(tmp_path)

        if not extracted_text.strip():
            await update.message.reply_text("Yaar is file mein kuch text nahi mila!")
            return

        extracted_text = extracted_text[:4000]
        prompt = f"Ye document ka content hai:\n\n{extracted_text}\n\nIs document ka summary de aur main points bata."
        ai_reply = get_ai_response(user_id, prompt)
        await update.message.reply_text(ai_reply)

    except Exception as e:
        await update.message.reply_text(f"Yaar kuch masla aa gaya: {str(e)}")

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    await update.message.chat.send_action("typing")
    try:
        await update.message.reply_text("🖼️ Image dekh raha hoon...")
        photo = update.message.photo[-1]
        file = await context.bot.get_file(photo.file_id)

        with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
            await file.download_to_drive(tmp.name)
            tmp_path = tmp.name

        with open(tmp_path, "rb") as img_file:
            image_data = base64.b64encode(img_file.read()).decode("utf-8")
        os.unlink(tmp_path)

        caption = update.message.caption or "Is image mein kya hai? Detail mein bata yaar."

        response = client.chat.completions.create(
            model="llama-3.2-11b-vision-preview",
            messages=[{
                "role": "user",
                "content": [
                    {"type": "text", "text": caption},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_data}"}}
                ]
            }]
        )
        ai_reply = response.choices[0].message.content
        await update.message.reply_text(ai_reply)

    except Exception as e:
        await update.message.reply_text(f"Yaar kuch masla aa gaya: {str(e)}")

def main():
    print("Bot chal raha hai...")
    app = ApplicationBuilder().token(TELEGRAM_BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("reset", reset))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_handler(MessageHandler(filters.VOICE, handle_voice))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    print("Bot ready!")
    app.run_polling()

if __name__ == "__main__":
    main()
